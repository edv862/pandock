import sys
import os
from openpyxl import load_workbook
from fpdf import FPDF


def excel_to_pdf(filename):
    data = read_excel(filename)
    print(data)
    """

    #creating a pdf in called test.pdf in the current directory
    pdf = FPDF()
    pdf.add_page()
    pdf.set_xy(0, 0)
    pdf.set_font('arial', 'B', 14)
    pdf.cell(60)
    pdf.cell(70, 10, 'Writing a PDF from python', 0, 2, 'C')
    pdf.cell(-40)
    pdf.cell(50, 10, 'Index Column', 1, 0, 'C')
    pdf.cell(40, 10, 'Col A', 1, 0, 'C')
    pdf.cell(40, 10, 'Col B', 1, 2, 'C')
    pdf.cell(-90)
    pdf.set_font('arial', '', 12)
    for i in range(0, len(df_2)-1):
        col_ind = str(i)
        col_a = str(df_2.A.ix[i])
        col_b = str(df_2.B.ix[i])
        pdf.cell(50, 10, '%s' % (col_ind), 1, 0, 'C')
        pdf.cell(40, 10, '%s' % (col_a), 0, 0, 'C')
        pdf.cell(40, 10, '%s' % (col_b), 0, 2, 'C')
        pdf.cell(-90)
    pdf.output('test.pdf', 'F')"""

# If verifies the availability of
def valid(value):
    if (value == None or value == 0 or value == False):
        return False
    return True

# Extracting data from the xls file
def read_excel(filename):
    work_book = load_workbook(filename)
    work_sheet_list = work_book.sheetnames
    ret = []

    # iterate over worksheets
    for workSheet in work_sheet_list:
        work_sheet = work_book[workSheet]
        line = []

        # iterate over rows
        for row in work_sheet['A{}:C{}'.format(work_sheet.min_row,work_sheet.max_row)]:
            # obtain lines 
            for cell in row:
                line.append(cell.value)

            available = line[2]
            # remove the third column
            line.remove(available)

            if valid(available):
                ret.append(line)
            line = []
    return ret

# Get file's extension
def get_extension(file):
    ext = os.path.splitext(file)[-1].lower()
    return ext

# to test in python
if __name__ == '__main__':
    # Take arguments
    args = []
    for arg in sys.argv[1:]:
        args.append(arg)

    # 1st argument is the file
    filename = args.pop(0)

    # Get file extension
    ext = get_extension(filename)

    text = ""

    if (ext == ".xls" or ext == ".xlsx"):
        excel_to_pdf(filename)
    else:
        # Invalid extension.
        pass
