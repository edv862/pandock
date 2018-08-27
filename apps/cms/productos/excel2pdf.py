import sys
import os
from pandock.settings.production import STATIC_ROOT
from openpyxl import load_workbook
from fpdf import FPDF


def excel_to_pdf(filename, title):
    data = read_excel(filename)

    # Creating the pdf instance and seting init values
    pdf = FPDF()
    pdf.add_page()
    pdf.set_xy(0, 0)
    background_url = os.path.join(STATIC_ROOT, 'img/pdf-lineas.png')
    #pdf.image(background_url, 0, 0, pdf.w, pdf.h + 20, type='png')

    # Generating The header of the file
    image_url = os.path.join(STATIC_ROOT, 'img/logo.png')
    pdf.image(image_url, x=10, y=8, w=75, h=27, type='png')
    pdf.set_font('arial', 'BU', 16)
    pdf.set_text_color(236, 126, 30)
    pdf.y = 40
    pdf.cell(20)
    pdf.cell(0, 10, title, 0, 2, 'C')

    # Generating the table header
    pdf.set_font('arial', 'B', 14)
    pdf.set_text_color(255)
    pdf.cell(150, 10, 'Productos', 1, 0, 'C', True)
    pdf.cell(20, 10, 'Unid', 1, 2, 'C', True)
    pdf.cell(-150)

    # Seting values of font for the table's values
    pdf.set_font('arial', '', 10)
    pdf.set_text_color(0)
    pdf.set_fill_color(255, 255, 255)

    for row in data:
        # Getting row's data
        height = 10
        col_a = str(row[0])
        col_b = str(row[1])
        # Save top coordinate
        top = pdf.y
        # Calculate x position of next cell
        offset = pdf.x + 150

        pdf.multi_cell(150, height, '%s' % (col_a), 1, 'J', True)

        # Reset y coordinate
        pdf.y = top
        # Move to computed offset
        pdf.x = offset

        if (len(col_a) > 73):
            height += height

        pdf.multi_cell(20, height, '%s' % (col_b), 1, 'J', True)
        pdf.cell(10)

        if pdf.y == 270:
            pdf.add_page()
            pdf.y = 40
            pdf.cell(10)
            pdf.image(image_url, x=10, y=8, w=75, h=27, type='png')
            pdf.set_font('arial', 'BU', 16)
            pdf.set_text_color(236, 126, 30)
            pdf.y = 40
            pdf.cell(0, 10, title, 0, 2, 'C')
            # Seting values of font for the table's values
            pdf.set_font('arial', '', 12)
            pdf.set_text_color(0)
            pdf.set_fill_color(255, 255, 255)
    return pdf


# If verifies the availability of
def valid(value):
    # if (value is None or value.lower() == "no"):
    if (value is None):
        return False
    return True


# Extracting data from the xls file
def read_excel(filename):
    work_book = load_workbook(filename)
    work_sheet_list = work_book.sheetnames
    ret = []

    # iterate over worksheets
    for workSheet in work_sheet_list:
        work_sheet = work_book[workSheet]
        line = []

        # iterate over rows
        for row in work_sheet['A{}:C{}'.format(
            work_sheet.min_row, work_sheet.max_row
        )]:
            # obtain lines
            for cell in row:
                line.append(cell.value)

            available = line[2]
            # remove the third column
            line.remove(available)

            if valid(available):
                ret.append(line)
            line = []
    return ret


# Get file's extension
def get_extension(file):
    ext = os.path.splitext(str(file))[-1].lower()
    return ext


# Create the pdf from the excel received
def create_pdf(filename, title):
    # Get file extension
    ext = get_extension(filename)

    if (ext == ".xls" or ext == ".xlsx"):
        return excel_to_pdf(filename, title)
    return None


# to test in python
if __name__ == '__main__':
    # Take arguments
    args = []
    for arg in sys.argv[1:]:
        args.append(arg)

    # 1st argument is the file
    filename = args.pop(0)

    # Get file extension
    ext = get_extension(filename)

    text = ""

    if (ext == ".xls" or ext == ".xlsx"):
        excel_to_pdf(filename, 'Los Pescaditos')
    else:
        # Invalid extension.
        pass
